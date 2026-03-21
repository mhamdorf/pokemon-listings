"""
create-input-workbook.py

Creates data/input/reference.xlsx with three sheets:
  - Sets: registry of all sets to process
  - Rarities: rarity rules used by the main pipeline
  - VariantOverrides: manual variant definitions for special sets

Run once to initialise. Re-running will overwrite the file.

Usage:
    uv run python scripts/create-input-workbook.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = os.path.join("data", "input", "reference.xlsx")

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
SUBHEADER_FILL = PatternFill("solid", start_color="BDD7EE")
SUBHEADER_FONT = Font(bold=True, name="Arial", size=10)
NORMAL_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="DCE6F1")
NOTE_FONT = Font(italic=True, color="888888", name="Arial", size=9)


def write_header(ws, headers, col_widths):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def write_row(ws, row_num, values, note_cols=None):
    fill = ALT_FILL if row_num % 2 == 0 else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = NOTE_FONT if note_cols and col in note_cols else NORMAL_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=(col == len(values)))
        if fill:
            cell.fill = fill


def create_sets_sheet(wb):
    ws = wb.create_sheet("Sets")
    headers = ["Set ID", "Set Name", "Series ID", "Series Name", "Set Type", "In Scope", "Notes"]
    col_widths = [14, 28, 12, 20, 12, 10, 45]
    write_header(ws, headers, col_widths)

    # Dropdown validation for set_type
    dv_type = DataValidation(type="list", formula1='"main,special"', allow_blank=False)
    dv_scope = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_scope)

    # Set type rules:
    #   - IDs ending in a letter suffix after .5 (e.g. sv10.5w) or containing .5 → special
    #   - IDs containing 'p' (promo sets) → main, In Scope = No
    #   - Everything else → main
    def derive_type(set_id):
        base = set_id.rstrip("abcdefghijklmnopqrstuvwxyz")
        return "special" if base.endswith(".5") else "main"

    def is_promo(set_id):
        return set_id.endswith("p") or set_id.startswith("mep") or set_id.startswith("svp")

    sets = [
        # Mega Evolution
        ("mep",    "MEP Black Star Promos",  "me", "Mega Evolution",   "main",    "Yes", ""),
        ("me01",   "Mega Evolution",          "me", "Mega Evolution",   "main",    "Yes", ""),
        ("me02",   "Phantasmal Flames",       "me", "Mega Evolution",   "main",    "Yes", ""),
        ("me02.5", "Ascended Heroes",         "me", "Mega Evolution",   "special", "Yes", "Named reverse holos — requires VariantOverrides before pipeline runs"),
        # Scarlet & Violet
        ("sv01",   "Scarlet & Violet",        "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("svp",    "SVP Black Star Promos",   "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv02",   "Paldea Evolved",          "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv03",   "Obsidian Flames",         "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv03.5", "151",                     "sv", "Scarlet & Violet", "special", "Yes", ""),
        ("sv04",   "Paradox Rift",            "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv04.5", "Paldean Fates",           "sv", "Scarlet & Violet", "special", "Yes", ""),
        ("sv05",   "Temporal Forces",         "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv06",   "Twilight Masquerade",     "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv06.5", "Shrouded Fable",          "sv", "Scarlet & Violet", "special", "Yes", ""),
        ("sv07",   "Stellar Crown",           "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv08",   "Surging Sparks",          "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv08.5", "Prismatic Evolutions",    "sv", "Scarlet & Violet", "special", "Yes", ""),
        ("sv09",   "Journey Together",        "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv10",   "Destined Rivals",         "sv", "Scarlet & Violet", "main",    "Yes", ""),
        ("sv10.5w","White Flare",             "sv", "Scarlet & Violet", "special", "Yes", ""),
        ("sv10.5b","Black Bolt",              "sv", "Scarlet & Violet", "special", "Yes", ""),
    ]

    # Verify derive_type matches what we've explicitly set (catches future additions)
    for row in sets:
        set_id, _, _, _, set_type, _, _ = row
        assert derive_type(set_id) == set_type or is_promo(set_id), \
            f"Set type mismatch for {set_id}: expected {derive_type(set_id)}, got {set_type}"

    for row_num, row in enumerate(sets, 2):
        write_row(ws, row_num, row)
        dv_type.add(ws.cell(row=row_num, column=5))
        dv_scope.add(ws.cell(row=row_num, column=6))


def create_rarities_sheet(wb):
    ws = wb.create_sheet("Rarities")
    headers = ["Rarity", "Base Finish", "Can Reverse Holo", "Notes"]
    col_widths = [30, 14, 18, 50]
    write_header(ws, headers, col_widths)

    dv_finish = DataValidation(type="list", formula1='"Normal,Holo"', allow_blank=False)
    dv_reverse = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_finish)
    ws.add_data_validation(dv_reverse)

    rarities = [
        # Modern series — reverse holo eligible
        ("Common",                    "Normal", "Yes", ""),
        ("Uncommon",                  "Normal", "Yes", ""),
        ("Rare",                      "Holo",   "Yes", "Older sets only"),
        ("Rare Holo",                 "Holo",   "Yes", "DP–BW era naming"),
        ("Holo Rare",                 "Holo",   "Yes", "SWSH era naming"),
        # No reverse holo
        ("Double Rare",               "Holo",   "No",  ""),
        ("Ultra Rare",                "Holo",   "No",  ""),
        ("Illustration Rare",         "Holo",   "No",  ""),
        ("Special Illustration Rare", "Holo",   "No",  ""),
        ("Hyper Rare",                "Holo",   "No",  ""),
        ("ACE SPEC Rare",             "Holo",   "No",  ""),
        ("Radiant Rare",              "Holo",   "No",  ""),
        ("Amazing Rare",              "Holo",   "No",  ""),
        ("Shiny Rare",                "Holo",   "No",  ""),
        ("Shiny Ultra Rare",          "Holo",   "No",  ""),
        ("Mega Hyper Rare",           "Holo",   "No",  "Mega Evolution series"),
        ("Mega Attack Rare",          "Holo",   "No",  "Mega Evolution series"),
        ("Secret Rare",               "Holo",   "No",  ""),
        ("Full Art Trainer",          "Holo",   "No",  ""),
        ("Classic Collection",        "Holo",   "No",  ""),
    ]

    for row_num, row in enumerate(rarities, 2):
        write_row(ws, row_num, row, note_cols={4})
        dv_finish.add(ws.cell(row=row_num, column=2))
        dv_reverse.add(ws.cell(row=row_num, column=3))

    # Add note below table
    note_row = len(rarities) + 3
    cell = ws.cell(row=note_row, column=1,
                   value="Note: Any rarity not listed here will default to Base Finish = Holo, Can Reverse Holo = No.")
    cell.font = NOTE_FONT


def create_overrides_sheet(wb):
    ws = wb.create_sheet("VariantOverrides")
    headers = ["Set ID", "Local ID", "Card Name", "Variants (pipe-separated)", "Reviewed By", "Reviewed Date", "Notes"]
    col_widths = [12, 10, 28, 60, 16, 16, 40]
    write_header(ws, headers, col_widths)

    # Example row (greyed out as a guide)
    example = ["me02.5", "001", "Erika's Oddish",
               "Normal|Reverse Holo (Fire Energy)|Reverse Holo (Friend Ball)",
               "hamdo", "2026-03-21", "Confirmed via physical card"]
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))


def main():
    os.makedirs(os.path.join("data", "input"), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    create_sets_sheet(wb)
    create_rarities_sheet(wb)
    create_overrides_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")
    print("Sheets: Sets, Rarities, VariantOverrides")


if __name__ == "__main__":
    main()
