"""
export-ebay-csv.py

Generates an eBay Seller Hub multi-variation listing CSV from ebay.xlsx stock data.

Produces one parent row (listing metadata) and one child row per in-stock variant.
Upload the output CSV via Seller Hub → Reports → Upload → Create listings.

Pricing logic:
  - If a row in the Stock sheet has a Price value, that price is used.
  - Otherwise the default for that finish type is used (--price-normal/holo/rh).

Usage:
    uv run python scripts/export-ebay-csv.py \\
        --set me02.5 \\
        --title "✅ PICK YOUR CARD ✅ ASCENDED HEROES — Pokémon TCG Singles NM FREE POST" \\
        --location "Melbourne, VIC" \\
        --price-normal 1.00 \\
        --price-holo 1.25 \\
        --price-rh 1.50
"""

import sys
import os
import csv
import argparse

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

# --- Paths ---
EBAY_PATH      = os.path.join("data", "input", "ebay.xlsx")
CATALOGUE_PATH = os.path.join("data", "output", "catalogue.xlsx")
OUTPUT_DIR     = os.path.join("data", "output")

# --- eBay constants (AU) ---
CATEGORY_ID    = "183454"    # CCG Individual Cards
CONDITION_ID   = "4000"      # Ungraded
CONDITION_DESC = "400010"    # Near Mint or Better
FORMAT         = "FixedPrice"
DURATION       = "GTC"
COUNTRY        = "AU"
BRAND          = "Nintendo/Creatures Inc."
LANGUAGE       = "English"
ACTION_HEADER  = "Action(SiteID=Australia|Country=AU|Currency=AUD|Version=1193|CC=UTF-8)"

DESCRIPTION_TEMPLATE = """\
<div style="font-family:Arial,sans-serif;max-width:620px;line-height:1.6">
  <h2 style="color:#333">{set_name} — Pokémon TCG Singles</h2>
  <p><strong>All cards are Near Mint or better</strong>, stored in a smoke-free environment.</p>
  <h3>How to order multiple cards</h3>
  <ol>
    <li>Select a card from the dropdown menu</li>
    <li>Click <em>Add to cart</em></li>
    <li>Repeat for each card you want</li>
    <li>Checkout once — one shipping fee covers everything</li>
  </ol>
  <h3>Packaging &amp; postage</h3>
  <p>FREE untracked letter post Australia-wide. Every card is sleeved and placed in a rigid top loader before being posted in a sturdy envelope.</p>
  <h3>Returns</h3>
  <p>No returns accepted. Please ask any questions before purchasing — happy to provide additional photos or condition details.</p>
</div>"""

CSV_COLUMNS = [
    ACTION_HEADER,
    "Title",
    "CategoryID",
    "ConditionID",
    "CD:40001",
    "Format",
    "Duration",
    "Description",
    "RelationshipDetails",
    "Relationship",
    "StartPrice",
    "Quantity",
    "P:UPC",
    "CustomLabel",
    "PicURL",
    "ShippingProfileName",
    "ReturnProfileName",
    "Country",
    "Location",
    "C:Brand",
    "C:Set",
    "C:Language",
    "C:Game",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_finish(variant_label: str) -> str:
    """Classify a variant label into normal / holo / rh."""
    if variant_label == "Normal":
        return "normal"
    elif variant_label == "Holo":
        return "holo"
    else:
        return "rh"


def make_dropdown_label(local_id: str, name: str, variant_label: str) -> str:
    """
    Build the buyer-facing dropdown label, e.g. '001 Erika's Oddish - Normal'.
    eBay enforces a 65-character maximum per variation value.
    """
    label = f"{local_id} {name} - {variant_label}"
    if len(label) > 65:
        overhead = len(f"{local_id}  - {variant_label}")
        name = name[: 65 - overhead].rstrip()
        label = f"{local_id} {name} - {variant_label}"
    return label


def image_url(api_base: str) -> str:
    return f"{api_base.rstrip('/')}/high.jpg"


def empty_row() -> dict:
    return {col: "" for col in CSV_COLUMNS}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate eBay multi-variation listing CSV.")
    parser.add_argument("--set",              required=True,
                        help="Set ID to export (e.g. me02.5)")
    parser.add_argument("--title",            required=True,
                        help="Listing title — max 80 characters")
    parser.add_argument("--location",         required=True,
                        help="Your suburb and state, e.g. 'Melbourne, VIC'")
    parser.add_argument("--price-normal",     type=float, default=1.00,
                        help="Default price for Normal cards (default: 1.00)")
    parser.add_argument("--price-holo",       type=float, default=1.25,
                        help="Default price for Holo cards (default: 1.25)")
    parser.add_argument("--price-rh",         type=float, default=1.50,
                        help="Default price for Reverse Holo cards (default: 1.50)")
    parser.add_argument("--shipping-profile", default="",
                        help="Seller Hub shipping policy name (optional)")
    parser.add_argument("--return-profile",   default="",
                        help="Seller Hub return policy name (optional)")
    args = parser.parse_args()

    set_id = args.set

    if len(args.title) > 80:
        print(f"[WARN] Title is {len(args.title)} chars — eBay max is 80. Truncating.")
        args.title = args.title[:80]

    default_prices = {
        "normal": args.price_normal,
        "holo":   args.price_holo,
        "rh":     args.price_rh,
    }

    # ------------------------------------------------------------------
    # Load Stock sheet
    # ------------------------------------------------------------------
    if not os.path.exists(EBAY_PATH):
        print(f"[ERROR] {EBAY_PATH} not found. Run build-stock-sheet.py first.")
        sys.exit(1)

    wb_ebay  = openpyxl.load_workbook(EBAY_PATH, data_only=True)
    ws_stock = wb_ebay["Stock"]

    sh = [cell.value for cell in ws_stock[1]]
    sc = {h: i for i, h in enumerate(sh)}
    has_price = "Price" in sc

    stock_rows = []
    for row in ws_stock.iter_rows(min_row=2, values_only=True):
        if row[sc["Set ID"]] != set_id:
            continue
        qty = row[sc["Qty"]] or 0
        if qty < 1:
            continue
        stock_rows.append(row)

    wb_ebay.close()

    if not stock_rows:
        print(f"[ERROR] No in-stock variants found for '{set_id}'. Fill in Qty values in ebay.xlsx.")
        sys.exit(1)

    # ------------------------------------------------------------------
    # Load image URLs + set name from catalogue
    # ------------------------------------------------------------------
    if not os.path.exists(CATALOGUE_PATH):
        print(f"[ERROR] {CATALOGUE_PATH} not found. Run build-catalogue.py first.")
        sys.exit(1)

    wb_cat  = openpyxl.load_workbook(CATALOGUE_PATH, read_only=True, data_only=True)
    ws_cards = wb_cat["Cards"]
    ws_sets  = wb_cat["Sets"]

    ch = [cell.value for cell in next(ws_cards.iter_rows(min_row=1, max_row=1))]
    cc = {h: i for i, h in enumerate(ch)}

    image_map = {}
    for row in ws_cards.iter_rows(min_row=2, values_only=True):
        if row[cc["Set ID"]] == set_id:
            card_id = row[cc["Card ID"]]
            api_base = row[cc["Image (API URL)"]] or ""
            if api_base:
                image_map[card_id] = image_url(api_base)

    set_name = set_id
    seth = [cell.value for cell in next(ws_sets.iter_rows(min_row=1, max_row=1))]
    for row in ws_sets.iter_rows(min_row=2, values_only=True):
        if row[0] == set_id:
            set_name = row[1] or set_id
            break

    wb_cat.close()

    # ------------------------------------------------------------------
    # Build variant list
    # ------------------------------------------------------------------
    variants = []
    for row in stock_rows:
        local_id_raw  = str(row[sc["Local ID"]] or "").zfill(3)
        name          = row[sc["Name"]]          or ""
        variant_label = row[sc["Variant Label"]] or ""
        qty           = int(row[sc["Qty"]]       or 0)
        variant_id    = row[sc["Variant ID"]]    or ""

        price_override = row[sc["Price"]] if has_price else None
        finish  = get_finish(variant_label)
        price   = float(price_override) if price_override else default_prices[finish]

        card_id   = f"{set_id}-{local_id_raw}"
        pic_url   = image_map.get(card_id, "")
        dropdown  = make_dropdown_label(local_id_raw, name, variant_label)

        variants.append({
            "variant_id":    variant_id,
            "dropdown":      dropdown,
            "finish":        finish,
            "qty":           qty,
            "price":         price,
            "pic_url":       pic_url,
        })

    print(f"[INFO] Set:      {set_name} ({set_id})")
    print(f"[INFO] Variants: {len(variants)} in stock")
    print(f"[INFO] Prices:   Normal ${default_prices['normal']:.2f} | "
          f"Holo ${default_prices['holo']:.2f} | RH ${default_prices['rh']:.2f}")

    # ------------------------------------------------------------------
    # Hero image: prefer a Holo, fall back to first variant
    # ------------------------------------------------------------------
    hero_pic = next((v["pic_url"] for v in variants if v["finish"] == "holo"), "")
    if not hero_pic:
        hero_pic = variants[0]["pic_url"] if variants else ""

    # ------------------------------------------------------------------
    # Build RelationshipDetails for parent row
    # ------------------------------------------------------------------
    all_labels = ";".join(v["dropdown"] for v in variants)
    parent_rel = f"Card={all_labels}"

    custom_label_base = set_id.upper().replace(".", "")

    # ------------------------------------------------------------------
    # Write CSV
    # ------------------------------------------------------------------
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, f"{set_id}-ebay-listing.csv")

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()

        # Parent row
        parent = empty_row()
        parent[ACTION_HEADER]        = "Add"
        parent["Title"]              = args.title
        parent["CategoryID"]         = CATEGORY_ID
        parent["ConditionID"]        = CONDITION_ID
        parent["CD:40001"]           = CONDITION_DESC
        parent["Format"]             = FORMAT
        parent["Duration"]           = DURATION
        parent["Description"]        = DESCRIPTION_TEMPLATE.format(set_name=set_name)
        parent["RelationshipDetails"] = parent_rel
        parent["CustomLabel"]        = f"{custom_label_base}-LISTING"
        parent["PicURL"]             = hero_pic
        parent["ShippingProfileName"] = args.shipping_profile
        parent["ReturnProfileName"]  = args.return_profile
        parent["Country"]            = COUNTRY
        parent["Location"]           = args.location
        parent["C:Brand"]            = BRAND
        parent["C:Set"]              = set_name
        parent["C:Language"]         = LANGUAGE
        parent["C:Game"]             = "Pokémon TCG"
        writer.writerow(parent)

        # Child rows — one per in-stock variant
        for v in variants:
            child = empty_row()
            child[ACTION_HEADER]         = "Add"
            child["CD:40001"]            = CONDITION_DESC
            child["RelationshipDetails"] = f"Card={v['dropdown']}"
            child["Relationship"]        = "Variation"
            child["StartPrice"]          = f"{v['price']:.2f}"
            child["Quantity"]            = str(v["qty"])
            child["P:UPC"]               = "Does not apply"
            child["CustomLabel"]         = v["variant_id"]
            child["PicURL"]              = v["pic_url"]
            writer.writerow(child)

    print(f"[DONE] {output_path}")
    print(f"       1 parent row + {len(variants)} variation rows")
    print()
    print("Upload: Seller Hub → Reports → Upload → Create listings")
    print("Note:   Add your shipping and return policies in Seller Hub after upload")
    print("        if you didn't supply --shipping-profile / --return-profile.")


if __name__ == "__main__":
    main()
