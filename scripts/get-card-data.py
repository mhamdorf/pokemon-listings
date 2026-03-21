import requests
import os
import re
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- Config ---
BASE_URL = "https://api.tcgdex.net/v2/en"
BULBAPEDIA_BASE = "https://bulbapedia.bulbagarden.net/wiki"
OUTPUT_DIR = "pokemon_cards"
IMG_SMALL_DIR = os.path.join(OUTPUT_DIR, "images_small")
IMG_LARGE_DIR = os.path.join(OUTPUT_DIR, "images_large")
EXCEL_PATH = os.path.join(OUTPUT_DIR, "pokemon_cards.xlsx")

# Single set to test with
TARGET_SET_ID = "me02.5"
SET_BULBAPEDIA_NAME = "Ascended_Heroes"  # Used to construct card page URLs

# Rarities that never get reverse holos
NO_REVERSE_RARITIES = {
    "Double Rare", "Ultra Rare", "Illustration Rare",
    "Special Illustration Rare", "Hyper Rare", "Mega Hyper Rare",
    "Mega Attack Rare", "Shiny Rare", "Shiny Ultra Rare", "Crown"
}

# Cache Bulbapedia results to avoid re-fetching
bulbapedia_cache = {}


def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name).strip().replace(" ", "_")


def fetch_json(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"    [ERROR] Failed to fetch {url}: {e}")
                return None


def fetch_html(url, retries=3):
    if url in bulbapedia_cache:
        return bulbapedia_cache[url]
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            bulbapedia_cache[url] = r.text
            time.sleep(0.3)  # Be polite to Bulbapedia
            return r.text
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"    [WARN] Bulbapedia fetch failed: {e}")
                return None


def download_image(url, filepath, retries=3):
    filename = os.path.basename(filepath)
    if os.path.exists(filepath):
        print(f"    [SKIP] Already exists: {filename}")
        return True
    print(f"    [DOWNLOAD] {filename}")
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=20, stream=True)
            r.raise_for_status()
            with open(filepath, "wb") as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            return True
        except Exception as e:
            if attempt < retries - 1:
                print(f"    [RETRY {attempt+2}] {filename}")
                time.sleep(2)
            else:
                print(f"    [ERROR] Failed to download: {filename} — {e}")
                return False


def get_bulbapedia_variants(card_name, set_name, card_number):
    """
    Scrape the individual card's Bulbapedia page and extract its
    reverse holo variant names from the description text.

    Bulbapedia consistently phrases it as:
      "...two Mirror Holofoil variants featuring either a X or a Y."
      or for single reverse:
      "...available in the regular Mega Evolution Series reverse pattern only."

    Returns a list of variant label strings, e.g.:
      ["Reverse Holo (Fire Energy)", "Reverse Holo (Friend Ball)"]
      ["Reverse Holo"]
      []   ← if no reverse
    """
    # Build Bulbapedia page title: e.g. "Charmander_(Ascended_Heroes_20)"
    safe_card = card_name.replace(" ", "_").replace("'", "%27")
    number_clean = card_number.lstrip("0") if card_number.isdigit() else card_number
    page_title = f"{safe_card}_({set_name}_{number_clean})"
    url = f"{BULBAPEDIA_BASE}/{page_title}"

    print(f"    [BULBAPEDIA] {url}")
    html = fetch_html(url)
    if not html:
        return None  # None = unknown, fall back gracefully

    soup = BeautifulSoup(html, "html.parser")

    # Get all paragraph text from the main content
    content = soup.find(id="mw-content-text")
    if not content:
        return None

    full_text = content.get_text(" ", strip=True)

    # Pattern 1: two variants — "featuring either a X or a Y."
    # Capture v1 up to optional symbol/logo/pattern + "or", v2 up to sentence end
    match = re.search(
        r"two Mirror Holofoil variants featuring either (?:a |an )?(.+?)(?:\s+(?:symbol|logo|pattern))?\s+or (?:a |an )?([^.]+?)\.",
        full_text, re.IGNORECASE
    )
    if match:
        # Strip trailing descriptor words from both groups
        def clean(s):
            return re.sub(r"\s+(symbol|logo|pattern)$", "", s.strip(), flags=re.IGNORECASE)
        v1 = clean(match.group(1))
        v2 = clean(match.group(2))
        return [f"Reverse Holo ({v1})", f"Reverse Holo ({v2})"]

    # Pattern 2: single reverse — standard pattern
    if re.search(r"reverse pattern only", full_text, re.IGNORECASE):
        return ["Reverse Holo"]

    # Pattern 3: explicitly no reverse (secret rares etc.)
    if re.search(r"no reverse|not available in reverse|does not have a reverse", full_text, re.IGNORECASE):
        return []

    return None  # Unknown — caller will fall back to TCGdex data


def get_cards_for_set(set_id):
    data = fetch_json(f"{BASE_URL}/sets/{set_id}")
    if not data:
        return [], "Unknown Set"
    return data.get("cards", []), data.get("name", set_id)


def get_card_detail(set_id, local_id):
    return fetch_json(f"{BASE_URL}/sets/{set_id}/{local_id}")


def build_variant_rows(card, card_set_id, set_name, bulbapedia_set_name):
    """
    Build all rows for a card including proper reverse variant names.
    Priority: Bulbapedia scrape > TCGdex variants fallback.
    """
    name = card.get("name", "Unknown")
    number = card.get("localId", "")
    rarity = card.get("rarity", "")
    artist = card.get("illustrator", "")
    image_base = card.get("image", "")
    tcgdex_variants = card.get("variants", {})
    rows = []

    # --- Determine normal/holo rows ---
    # Ignore TCGdex holo flag as it's unreliable for this set.
    # Holo only exists on Rare rarity cards in Ascended Heroes.
    has_first_ed = tcgdex_variants.get("firstEdition", False)
    has_promo = tcgdex_variants.get("wPromo", False)
    has_reverse = tcgdex_variants.get("reverse", False)
    is_rare = (rarity == "Rare")

    # Rares: base card is Holo (no Normal). Common/Uncommon: base card is Normal.
    base_label = "Holo" if is_rare else "Normal"
    rows.append(base_label)

    for flag, label in [
        (has_first_ed, "First Edition"),
        (has_promo, "Promo"),
    ]:
        if flag:
            rows.append(label)

    # --- Determine reverse variants ---
    if has_reverse and rarity not in NO_REVERSE_RARITIES:
        # Try Bulbapedia first
        bp_variants = get_bulbapedia_variants(name, bulbapedia_set_name, number)

        if bp_variants is not None:
            rows.extend(bp_variants)
            print(f"    [VARIANTS] Bulbapedia: {bp_variants if bp_variants else 'no reverse'}")
        else:
            # Fallback: generic Reverse Holo
            rows.append("Reverse Holo")
            print(f"    [VARIANTS] Bulbapedia unavailable, using generic Reverse Holo")
    # --- Build row dicts with image downloads ---
    result = []
    for variant_label in rows:
        variant_key = sanitize_filename(variant_label)
        safe_name = sanitize_filename(f"{card_set_id}_{number}_{name}_{variant_key}")
        img_small = f"{safe_name}_small.jpg"
        img_large = f"{safe_name}_large.jpg"

        if image_base:
            download_image(f"{image_base}/low.jpg", os.path.join(IMG_SMALL_DIR, img_small))
            download_image(f"{image_base}/high.jpg", os.path.join(IMG_LARGE_DIR, img_large))
        else:
            img_small = img_large = "N/A"

        result.append({
            "set_name": set_name,
            "set_id": card_set_id,
            "number": number,
            "name": name,
            "variant": variant_label,
            "rarity": rarity,
            "artist": artist,
            "img_small": img_small,
            "img_large": img_large,
        })

    return result


def setup_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cards"

    headers = ["Set Name", "Set ID", "Card Number", "Name", "Variant", "Rarity", "Artist", "Image (Small)", "Image (Large)"]
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    col_widths = [25, 10, 14, 28, 30, 22, 22, 38, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    return wb, ws


def write_row(ws, row_num, data):
    values = [
        data["set_name"], data["set_id"], data["number"], data["name"],
        data["variant"], data["rarity"], data["artist"],
        data["img_small"], data["img_large"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
        if row_num % 2 == 0:
            cell.fill = PatternFill("solid", start_color="DCE6F1")


def main():
    os.makedirs(IMG_SMALL_DIR, exist_ok=True)
    os.makedirs(IMG_LARGE_DIR, exist_ok=True)

    print(f"Fetching set: {TARGET_SET_ID}")
    card_stubs, set_name = get_cards_for_set(TARGET_SET_ID)
    print(f"  Set name: {set_name}")
    print(f"  {len(card_stubs)} cards found\n")

    wb, ws = setup_excel()
    excel_row = 2
    total_rows = 0
    errors = []

    for i, stub in enumerate(card_stubs, 1):
        local_id = stub.get("localId")
        card_full_id = stub.get("id", "")
        card_set_id = card_full_id.rsplit("-", 1)[0] if "-" in card_full_id else TARGET_SET_ID
        card_name = stub.get("name", local_id)

        print(f"\n  [{i}/{len(card_stubs)}] {card_name} ({card_set_id}-{local_id})")

        card = get_card_detail(card_set_id, local_id)
        if not card:
            print(f"    [ERROR] Could not fetch card data")
            errors.append(f"{card_set_id}/{local_id}")
            continue

        variant_rows = build_variant_rows(card, card_set_id, set_name, SET_BULBAPEDIA_NAME)
        print(f"    → {len(variant_rows)} row(s): {', '.join(r['variant'] for r in variant_rows)}")

        for row_data in variant_rows:
            write_row(ws, excel_row, row_data)
            excel_row += 1
            total_rows += 1

        time.sleep(0.1)

    wb.save(EXCEL_PATH)

    print(f"\n{'='*50}")
    print(f"Done! {len(card_stubs)} cards -> {total_rows} rows (including variants)")
    print(f"Excel:        {EXCEL_PATH}")
    print(f"Small images: {IMG_SMALL_DIR}/")
    print(f"Large images: {IMG_LARGE_DIR}/")
    if errors:
        print(f"\n{len(errors)} cards failed:")
        for e in errors:
            print(f"  - {e}")


if __name__ == "__main__":
    main()