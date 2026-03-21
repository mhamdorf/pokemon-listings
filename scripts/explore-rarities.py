"""
explore-rarities.py

Fetches all distinct rarities from TCGdex in one call, then samples a few
cards per rarity to determine their variant flag patterns.

Output: data/output/rarities_exploration.xlsx

Usage:
    uv run python scripts/explore-rarities.py
    uv run python scripts/explore-rarities.py --sample 5
"""

import requests
import argparse
import time
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "https://api.tcgdex.net/v2/en"
OUTPUT_PATH = os.path.join("data", "output", "rarities_exploration.xlsx")

EXCLUDED_SERIES = {"tcgp"}  # TCG Pocket — digital only, not physical cards

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL = PatternFill("solid", start_color="DCE6F1")
NORMAL_FONT = Font(name="Arial", size=10)


def fetch(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"  [ERROR] {url}: {e}")
                return None


def write_sheet_rarities(wb, rarity_data):
    ws = wb.create_sheet("Rarities Summary")
    headers = [
        "Rarity",
        "Cards Sampled",
        "normal=T / holo=F / reverse=F",
        "normal=T / holo=T / reverse=F",
        "normal=T / holo=F / reverse=T",
        "normal=T / holo=T / reverse=T",
        "normal=F / holo=T / reverse=F",
        "normal=F / holo=T / reverse=T",
        "normal=F / holo=F / reverse=F",
        "Other Combinations",
        "Example Sets",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    col_widths = [28, 14, 22, 22, 22, 22, 22, 22, 22, 18, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    known_combos = [
        (True, False, False),
        (True, True, False),
        (True, False, True),
        (True, True, True),
        (False, True, False),
        (False, True, True),
        (False, False, False),
    ]

    for row_num, (rarity, data) in enumerate(sorted(rarity_data.items()), 2):
        combo_counts = data["combos"]
        total = data["total"]
        sets = sorted(data["sets"])[:5]  # show up to 5 example sets

        other = sum(
            count for combo, count in combo_counts.items()
            if combo not in known_combos
        )

        values = [
            rarity,
            total,
            *[combo_counts.get(c, 0) for c in known_combos],
            other,
            ", ".join(sets),
        ]

        fill = ALT_FILL if row_num % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"


def write_sheet_raw(wb, raw_rows):
    ws = wb.create_sheet("Raw Sample")
    headers = ["Rarity", "Card ID", "Card Name", "Set Name",
               "variant.normal", "variant.holo", "variant.reverse",
               "variant.firstEdition", "variant.wPromo"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [28, 16, 30, 28, 14, 12, 14, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row_num, row in enumerate(raw_rows, 2):
        fill = ALT_FILL if row_num % 2 == 0 else None
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sample", type=int, default=3,
                        help="Cards to sample per rarity (default 3)")
    args = parser.parse_args()

    print("Fetching all rarities...")
    rarities = fetch(f"{BASE_URL}/rarities")
    if not rarities:
        print("Failed to fetch rarities.")
        return

    print("Fetching excluded series set IDs...")
    excluded_set_ids = set()
    for series_id in EXCLUDED_SERIES:
        series_data = fetch(f"{BASE_URL}/series/{series_id}")
        if series_data:
            for s in series_data.get("sets", []):
                excluded_set_ids.add(s.get("id", ""))
    print(f"Excluding {len(excluded_set_ids)} sets from: {', '.join(EXCLUDED_SERIES)}\n")

    print(f"Found {len(rarities)} rarities. Sampling {args.sample} cards each...\n")

    rarity_data = defaultdict(lambda: {"total": 0, "combos": defaultdict(int), "sets": set()})
    raw_rows = []

    for i, rarity_name in enumerate(rarities, 1):
        print(f"  [{i}/{len(rarities)}] {rarity_name}")

        rarity_detail = fetch(f"{BASE_URL}/rarities/{rarity_name.replace(' ', '%20')}")
        if not rarity_detail:
            continue

        cards = rarity_detail.get("cards", [])[:args.sample]

        for card_stub in cards:
            card_id = card_stub.get("id", "")
            # card id format: setId-localId, skip malformed or excluded stubs
            if "-" not in card_id or "?" in card_id:
                continue
            set_id, local_id = card_id.rsplit("-", 1)
            if set_id in excluded_set_ids:
                continue

            card = fetch(f"{BASE_URL}/sets/{set_id}/{local_id}")
            if not card:
                continue

            variants = card.get("variants", {})
            normal = variants.get("normal", False)
            holo = variants.get("holo", False)
            reverse = variants.get("reverse", False)
            first_ed = variants.get("firstEdition", False)
            wpromo = variants.get("wPromo", False)
            set_name = card.get("set", {}).get("name", set_id)

            combo = (normal, holo, reverse)
            rarity_data[rarity_name]["total"] += 1
            rarity_data[rarity_name]["combos"][combo] += 1
            rarity_data[rarity_name]["sets"].add(set_name)

            raw_rows.append([
                rarity_name, card_id, card.get("name"),
                set_name, normal, holo, reverse, first_ed, wpromo
            ])

            time.sleep(0.05)

    os.makedirs(os.path.join("data", "output"), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet_rarities(wb, rarity_data)
    write_sheet_raw(wb, raw_rows)

    wb.save(OUTPUT_PATH)

    print(f"\nDone. {len(rarity_data)} rarities sampled.")
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
