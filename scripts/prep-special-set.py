"""
prep-special-set.py

One-time prep tool for special sets. Fetches all cards for a set,
attempts to scrape variant names from Bulbapedia, and writes a draft
to the VariantOverrides sheet in data/input/reference.xlsx.

The human must then review each row, correct any wrong/missing variant
names, and fill in 'reviewed_by' and 'reviewed_date' before the main
pipeline will process the set.

Already-reviewed entries (reviewed_by + reviewed_date filled) are never
overwritten — re-running is safe.

Usage:
    uv run python scripts/prep-special-set.py --set me02.5 --bulbapedia-name Ascended_Heroes

    # Without Bulbapedia (uses TCGdex flags only for base guess)
    uv run python scripts/prep-special-set.py --set me02.5
"""

import requests
import argparse
import re
import time
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_PATH       = os.path.join("data", "input", "reference.xlsx")
TCGDEX_URL       = "https://api.tcgdex.net/v2/en"
BULBAPEDIA_BASE  = "https://bulbapedia.bulbagarden.net/wiki"

DRAFT_FILL     = PatternFill("solid", start_color="FFF2CC")   # yellow — needs review
REVIEWED_FILL  = PatternFill("solid", start_color="E2EFDA")   # green — reviewed
NORMAL_FONT    = Font(name="Arial", size=10)
NOTE_FONT      = Font(italic=True, color="888888", name="Arial", size=9)

html_cache = {}


# ---------------------------------------------------------------------------
# Fetch helpers
# ---------------------------------------------------------------------------

def fetch_json(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"    [ERROR] {url}: {e}")
                return None


def fetch_html(url, retries=3):
    if url in html_cache:
        return html_cache[url]
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            html_cache[url] = r.text
            time.sleep(0.4)  # be polite to Bulbapedia
            return r.text
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"    [WARN] Bulbapedia fetch failed: {e}")
                return None


# ---------------------------------------------------------------------------
# Bulbapedia scraping
# ---------------------------------------------------------------------------

def scrape_bulbapedia_variants(card_name, bulbapedia_set_name, local_id):
    """
    Attempt to scrape reverse holo variant names from the card's Bulbapedia page.

    Returns:
        list  — variant name strings found (may be empty if no reverse)
        None  — could not determine (page not found, pattern not matched)
    """
    safe_name    = card_name.replace(" ", "_").replace("'", "%27")
    number_clean = local_id.lstrip("0") if local_id.isdigit() else local_id
    page_title   = f"{safe_name}_({bulbapedia_set_name}_{number_clean})"
    url          = f"{BULBAPEDIA_BASE}/{page_title}"

    html = fetch_html(url)
    if not html:
        return None

    soup    = BeautifulSoup(html, "html.parser")
    content = soup.find(id="mw-content-text")
    if not content:
        return None

    text = content.get_text(" ", strip=True)

    # Two named variants
    match = re.search(
        r"two Mirror Holofoil variants featuring either (?:a |an )?(.+?)"
        r"(?:\s+(?:symbol|logo|pattern))?\s+or (?:a |an )?([^.]+?)\.",
        text, re.IGNORECASE
    )
    if match:
        def clean(s):
            return re.sub(r"\s+(symbol|logo|pattern)$", "", s.strip(), flags=re.IGNORECASE)
        return [f"Reverse Holo ({clean(match.group(1))})",
                f"Reverse Holo ({clean(match.group(2))})"]

    # Single standard reverse
    if re.search(r"reverse pattern only", text, re.IGNORECASE):
        return ["Reverse Holo"]

    # Explicitly no reverse
    if re.search(r"no reverse|not available in reverse|does not have a reverse",
                 text, re.IGNORECASE):
        return []

    return None  # inconclusive


# ---------------------------------------------------------------------------
# Variant guessing from TCGdex flags + rarity
# ---------------------------------------------------------------------------

def guess_variants_from_tcgdex(card, rarities):
    """
    Build a best-guess variant list from TCGdex flags and rarity rules.
    Used as fallback when Bulbapedia scraping is unavailable or inconclusive.
    """
    rarity   = card.get("rarity", "") or ""
    tcg_vars = card.get("variants", {})

    rarity_rule = rarities.get(rarity, {"base_finish": "Holo", "can_reverse_holo": False})
    base_finish = rarity_rule["base_finish"]
    can_reverse = rarity_rule["can_reverse_holo"]

    variants = [base_finish]

    if tcg_vars.get("firstEdition"):
        variants.append("First Edition")
    if tcg_vars.get("wPromo"):
        variants.append("Promo")
    if can_reverse and tcg_vars.get("reverse"):
        # For special sets we can't name the reverse variants without Bulbapedia
        # Use a placeholder the human can correct
        variants.append("Reverse Holo (???)")

    return variants


# ---------------------------------------------------------------------------
# Load reference workbook data
# ---------------------------------------------------------------------------

def load_rarities(wb):
    ws = wb["Rarities"]
    rarities = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rarity, base_finish, can_reverse, *_ = row
        rarities[rarity] = {
            "base_finish": base_finish,
            "can_reverse_holo": (can_reverse == "Yes"),
        }
    return rarities


def load_existing_overrides(ws_overrides):
    """
    Returns a dict of (set_id, local_id) -> row_number for existing entries,
    and a set of (set_id, local_id) that are already reviewed.
    """
    existing     = {}  # (set_id, local_id) -> row_num
    reviewed     = set()
    for row_num, row in enumerate(ws_overrides.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue
        set_id, local_id = row[0], str(row[1])
        existing[(set_id, local_id)] = row_num
        if row[4] and row[5]:  # reviewed_by and reviewed_date both filled
            reviewed.add((set_id, local_id))
    return existing, reviewed


# ---------------------------------------------------------------------------
# Write to VariantOverrides sheet
# ---------------------------------------------------------------------------

def write_override_row(ws, row_num, set_id, local_id, card_name,
                       variants, source, is_reviewed=False):
    fill = REVIEWED_FILL if is_reviewed else DRAFT_FILL
    values = [
        set_id, local_id, card_name,
        "|".join(variants),
        "",    # reviewed_by — human fills this
        "",    # reviewed_date — human fills this
        f"Draft source: {source}",
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
        cell.fill = fill


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--set", required=True, help="Set ID to prepare (e.g. me02.5)")
    parser.add_argument("--bulbapedia-name", default=None,
                        help="Bulbapedia set name for URL construction (e.g. Ascended_Heroes). "
                             "If omitted, Bulbapedia scraping is skipped.")
    args = parser.parse_args()

    set_id          = args.set
    bulbapedia_name = args.bulbapedia_name

    if not os.path.exists(INPUT_PATH):
        print(f"Input workbook not found: {INPUT_PATH}")
        print("Run scripts/create-input-workbook.py first.")
        return

    # Load workbook
    wb             = load_workbook(INPUT_PATH)
    rarities       = load_rarities(wb)
    ws_overrides   = wb["VariantOverrides"]
    existing, reviewed = load_existing_overrides(ws_overrides)

    # Validate set exists and is special
    ws_sets = wb["Sets"]
    set_info = None
    for row in ws_sets.iter_rows(min_row=2, values_only=True):
        if row[0] == set_id:
            set_info = {"set_id": row[0], "set_name": row[1], "set_type": row[4]}
            break

    if not set_info:
        print(f"Set '{set_id}' not found in Sets sheet.")
        return
    if set_info["set_type"] != "special":
        print(f"Set '{set_id}' is type '{set_info['set_type']}' — only special sets need overrides.")
        return

    print(f"Preparing override sheet for: {set_info['set_name']} ({set_id})")
    if bulbapedia_name:
        print(f"Bulbapedia set name: {bulbapedia_name}")
    else:
        print("No --bulbapedia-name provided — Bulbapedia scraping skipped, using TCGdex flags only.")
    print()

    # Fetch all cards for the set
    set_data = fetch_json(f"{TCGDEX_URL}/sets/{set_id}")
    if not set_data:
        print("Failed to fetch set data.")
        return

    card_stubs  = set_data.get("cards", [])
    total       = len(card_stubs)
    next_row    = max(existing.values(), default=1) + 1 if existing else 2

    stats = {"skipped": 0, "updated": 0, "new": 0, "errors": 0}

    for i, stub in enumerate(card_stubs, 1):
        local_id  = str(stub.get("localId", ""))
        card_name = stub.get("name", local_id)
        key       = (set_id, local_id)

        print(f"  [{i}/{total}] {card_name} ({set_id}-{local_id})", end="")

        # Skip already-reviewed entries
        if key in reviewed:
            print(f" [SKIP — already reviewed]")
            stats["skipped"] += 1
            continue

        # Fetch full card data
        card = fetch_json(f"{TCGDEX_URL}/sets/{set_id}/{local_id}")
        if not card:
            print(f" [ERROR — could not fetch card]")
            stats["errors"] += 1
            continue

        # Try Bulbapedia first if name provided
        source   = "TCGdex flags"
        variants = None

        if bulbapedia_name:
            variants = scrape_bulbapedia_variants(card_name, bulbapedia_name, local_id)
            if variants is not None:
                # Prepend the base finish (Bulbapedia only tells us about reverse variants)
                rarity      = card.get("rarity", "") or ""
                rarity_rule = rarities.get(rarity, {"base_finish": "Holo"})
                base_finish = rarity_rule["base_finish"]
                variants    = [base_finish] + variants
                source      = "Bulbapedia"
            else:
                print(f" [Bulbapedia inconclusive, falling back to TCGdex]", end="")

        if variants is None:
            variants = guess_variants_from_tcgdex(card, rarities)

        print(f" -> {' | '.join(variants)} [{source}]")

        # Write or update row
        if key in existing:
            row_num = existing[key]
            stats["updated"] += 1
        else:
            row_num = next_row
            next_row += 1
            stats["new"] += 1

        write_override_row(ws_overrides, row_num, set_id, local_id,
                           card_name, variants, source)

        time.sleep(0.05)

    wb.save(INPUT_PATH)

    print(f"\n{'='*60}")
    print(f"Done: {total} cards processed")
    print(f"  New entries:     {stats['new']}")
    print(f"  Updated entries: {stats['updated']}")
    print(f"  Already reviewed (skipped): {stats['skipped']}")
    print(f"  Errors:          {stats['errors']}")
    print(f"\nNext steps:")
    print(f"  1. Open {INPUT_PATH}")
    print(f"  2. Review the VariantOverrides sheet (yellow rows need attention)")
    print(f"     - Fix any '(???)' placeholders")
    print(f"     - Verify Bulbapedia-sourced names against physical cards")
    print(f"  3. Fill in 'reviewed_by' and 'reviewed_date' for each confirmed row")
    print(f"  4. Run build-catalogue.py --set {set_id}")


if __name__ == "__main__":
    main()
