"""
build-stock-sheet.py

Seeds (or updates) the Stock sheet in data/input/ebay.xlsx from the master
catalogue. One row per card variant, with a Qty column for the user to fill in.

Run this once per set when adding new cards to ebay.xlsx. Existing Qty values
are preserved — only new rows are added.

Usage:
    uv run python scripts/build-stock-sheet.py --set me02.5
    uv run python scripts/build-stock-sheet.py --set me02.5 --max-card 217
"""

import sys
import os
import argparse

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

# --- Paths ---
CATALOGUE_PATH = os.path.join("data", "output", "catalogue.xlsx")
EBAY_PATH      = os.path.join("data", "input", "ebay.xlsx")

# --- Styling ---
ACCENT_COLOR      = "5C5BDB"
HEADER_FILL       = PatternFill("solid", start_color=ACCENT_COLOR)
HEADER_FONT       = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL          = PatternFill("solid", start_color="F4F3FF")
REF_FILL          = PatternFill("solid", start_color="F0F0F0")   # light grey — read-only reference cols
REF_FILL_ALT      = PatternFill("solid", start_color="E8E8F0")   # alt row variant
NORMAL_FONT       = Font(name="Calibri", size=10)
DIM_FONT          = Font(name="Calibri", size=10, color="888888")
INPUT_FILL        = PatternFill("solid", start_color="FFFDE7")   # soft yellow — user edits this
HEADER_ROW_HEIGHT = 26
BODY_ROW_HEIGHT   = 16

# Columns: (header, width, is_user_input)
COLUMNS = [
    ("Variant ID",    34, False),
    ("Set ID",        10, False),
    ("Local ID",      10, False),
    ("Name",          28, False),
    ("Variant Label", 34, False),
    ("Qty",            8, True),
]

QTY_COL_IDX = 6   # 1-based


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_variant_id(card_id: str, variant_label: str) -> str:
    """Must match the Variant ID formula in the Power Query Listing view."""
    return card_id + "-" + variant_label.lower().replace(" ", "-")


def variant_sort_key(variant: str) -> tuple:
    """Normal → Holo → Reverse Holo (alphabetical within group)."""
    if variant == "Normal":
        return (0, "")
    elif variant == "Holo":
        return (1, "")
    elif variant.startswith("Reverse"):
        return (2, variant)
    else:
        return (3, variant)


def load_existing_qtys(ws) -> dict:
    """Read existing Variant ID → Qty mapping from the Stock sheet."""
    qtys = {}
    headers = [cell.value for cell in ws[1]]
    try:
        vid_idx = headers.index("Variant ID")
        qty_idx = headers.index("Qty")
    except ValueError:
        return qtys
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid = row[vid_idx]
        qty = row[qty_idx]
        if vid:
            qtys[vid] = qty if qty is not None else 0
    return qtys


def write_header(ws):
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for col_idx, (col_name, col_width, _) in enumerate(COLUMNS, 1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"


def write_data_row(ws, row_idx: int, values: list, existing_qty: int):
    is_alt = (row_idx % 2 == 0)

    for col_idx, ((_, __, is_input), value) in enumerate(zip(COLUMNS, values), 1):
        if col_idx == QTY_COL_IDX:
            cell           = ws.cell(row=row_idx, column=col_idx, value=existing_qty)
            cell.fill      = INPUT_FILL
            cell.font      = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", horizontal="center")
        else:
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DIM_FONT
            cell.fill      = REF_FILL_ALT if is_alt else REF_FILL
            cell.alignment = Alignment(vertical="center", indent=1)

    ws.row_dimensions[row_idx].height = BODY_ROW_HEIGHT


def add_table(ws, last_row: int):
    last_col = get_column_letter(len(COLUMNS))
    table = Table(displayName="Stock", ref=f"A1:{last_col}{last_row}")
    ws.add_table(table)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Seed Stock sheet in ebay.xlsx from the master catalogue.")
    parser.add_argument("--set",      required=True, help="Set ID to add (e.g. me02.5)")
    parser.add_argument("--max-card", type=int, default=None,
                        help="Only include cards with a numeric Local ID <= this value (e.g. 217)")
    args = parser.parse_args()

    set_id   = args.set
    max_card = args.max_card

    # ------------------------------------------------------------------
    # Load catalogue
    # ------------------------------------------------------------------
    if not os.path.exists(CATALOGUE_PATH):
        print(f"[ERROR] Catalogue not found at {CATALOGUE_PATH}. Run build-catalogue.py first.")
        sys.exit(1)

    wb_cat = openpyxl.load_workbook(CATALOGUE_PATH, read_only=True, data_only=True)
    ws_cat = wb_cat["Cards"]

    headers = [cell.value for cell in next(ws_cat.iter_rows(min_row=1, max_row=1))]
    col     = {h: i for i, h in enumerate(headers)}

    # ------------------------------------------------------------------
    # Read and expand variants from catalogue
    # ------------------------------------------------------------------
    new_rows = []
    skipped  = 0

    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        if row[col["Set ID"]] != set_id:
            continue

        local_id_raw = row[col["Local ID"]]
        try:
            local_id_int = int(str(local_id_raw).strip())
        except (ValueError, TypeError):
            local_id_int = 9999

        if max_card is not None and local_id_int > max_card:
            skipped += 1
            continue

        card_id      = row[col["Card ID"]]    or ""
        name         = row[col["Name"]]       or ""
        variants_raw = row[col["Variants"]]   or ""
        variants     = [v.strip() for v in str(variants_raw).split("|") if v.strip()]

        for variant in sorted(variants, key=variant_sort_key):
            variant_id = make_variant_id(card_id, variant)
            new_rows.append((local_id_int, local_id_raw, card_id, name, variant, variant_id))

    wb_cat.close()

    if not new_rows:
        print(f"[ERROR] No cards found for set '{set_id}' in the catalogue.")
        sys.exit(1)

    new_rows.sort(key=lambda r: (r[0], variant_sort_key(r[4])))
    print(f"[INFO] {len(new_rows)} card+variant rows found for {set_id}" +
          (f" (cards 001–{max_card:03d}, {skipped} above limit skipped)" if max_card else ""))

    # ------------------------------------------------------------------
    # Load or create ebay.xlsx, preserve existing Qtys
    # ------------------------------------------------------------------
    if os.path.exists(EBAY_PATH):
        wb = openpyxl.load_workbook(EBAY_PATH)
        if "Stock" in wb.sheetnames:
            existing_qtys = load_existing_qtys(wb["Stock"])
            print(f"[INFO] Loaded {len(existing_qtys)} existing Qty entries from ebay.xlsx")
            del wb["Stock"]
        else:
            existing_qtys = {}
        ws = wb.create_sheet("Stock", 0)
    else:
        existing_qtys = {}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock"

    new_count      = 0
    preserved_count = 0

    write_header(ws)

    for row_idx, (local_id_int, local_id_raw, card_id, name, variant, variant_id) in enumerate(new_rows, 2):
        existing_qty = existing_qtys.get(variant_id, 0)
        if variant_id in existing_qtys:
            preserved_count += 1
        else:
            new_count += 1

        values = [variant_id, set_id, local_id_raw, name, variant, existing_qty]
        write_data_row(ws, row_idx, values, existing_qty)

    add_table(ws, last_row=len(new_rows) + 1)

    os.makedirs(os.path.dirname(EBAY_PATH), exist_ok=True)
    wb.save(EBAY_PATH)

    print(f"[DONE] ebay.xlsx updated → {EBAY_PATH}")
    print(f"       {preserved_count} existing Qty values preserved, {new_count} new rows added at Qty=0")
    print()
    print("Next step: open ebay.xlsx, fill in the Qty column (yellow) for cards you have.")
    print("Then add the Listing Power Query view — see docs/how-to/create-listing-view.md")


if __name__ == "__main__":
    main()
