"""
explore-card-schema.py

Fetches a single card from TCGdex and writes every attribute to an Excel
workbook so you can see what's available to include in the final output.

Usage:
    uv run python scripts/explore-card-schema.py
    uv run python scripts/explore-card-schema.py --set sv01 --card 1

Output: pokemon_cards/card_schema.xlsx
"""

import requests
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "https://api.tcgdex.net/v2/en"
DEFAULT_SET = "me02.5"
DEFAULT_CARD = "1"
OUTPUT_PATH = os.path.join("data", "output", "card_schema.xlsx")

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
SECTION_FILL = PatternFill("solid", start_color="BDD7EE")
SECTION_FONT = Font(bold=True, name="Arial", size=10)
NORMAL_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="DCE6F1")


def fetch(url):
    r = requests.get(url, timeout=15)
    r.raise_for_status()
    return r.json()


def flatten(data, prefix=""):
    """Recursively flatten a nested dict/list into (field_path, value) pairs."""
    rows = []
    if isinstance(data, dict):
        for key, value in data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(value, (dict, list)):
                rows.extend(flatten(value, full_key))
            else:
                rows.append((full_key, value))
    elif isinstance(data, list):
        if not data:
            rows.append((prefix, "(empty list)"))
        else:
            for i, item in enumerate(data):
                rows.extend(flatten(item, f"{prefix}[{i}]"))
    else:
        rows.append((prefix, data))
    return rows


def write_sheet(wb, title, data, source_url):
    ws = wb.create_sheet(title=title)

    # Source URL row
    ws.append([f"Source: {source_url}"])
    ws.cell(row=1, column=1).font = Font(italic=True, name="Arial", size=9, color="666666")
    ws.merge_cells("A1:C1")
    ws.append([])

    # Headers
    header_row = 3
    for col, label in enumerate(["Field Path", "Value", "Type"], 1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18

    rows = flatten(data)
    prev_top = None
    data_row = header_row + 1

    for field, value in rows:
        top_level = field.split(".")[0].split("[")[0]
        is_section_start = top_level != prev_top
        prev_top = top_level

        fill = SECTION_FILL if is_section_start else (ALT_FILL if data_row % 2 == 0 else None)
        font = SECTION_FONT if is_section_start else NORMAL_FONT

        for col, val in enumerate([field, value, type(value).__name__], 1):
            cell = ws.cell(row=data_row, column=col, value=str(val) if val is not None else "")
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 2))
            if fill:
                cell.fill = fill

        ws.row_dimensions[data_row].height = 15
        data_row += 1

    ws.freeze_panes = "A4"
    return ws


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--set", default=DEFAULT_SET)
    parser.add_argument("--card", default=DEFAULT_CARD)
    args = parser.parse_args()

    set_url = f"{BASE_URL}/sets/{args.set}"
    card_url = f"{BASE_URL}/sets/{args.set}/{args.card}"

    print(f"Fetching set: {set_url}")
    set_data = fetch(set_url)

    print(f"Fetching card: {card_url}")
    card_data = fetch(card_url)

    os.makedirs(os.path.join("data", "output"), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_sheet(wb, "Card Attributes", card_data, card_url)

    # For set, strip the full card list — just show metadata + first card stub
    set_preview = {k: v for k, v in set_data.items() if k != "cards"}
    set_preview["cards (first 3)"] = set_data.get("cards", [])[:3]
    write_sheet(wb, "Set Attributes", set_preview, set_url)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
