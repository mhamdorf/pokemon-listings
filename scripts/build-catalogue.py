"""
build-catalogue.py

Main pipeline. Reads data/input/reference.xlsx for set configuration,
fetches card data from TCGdex, and writes the master catalogue to
data/output/catalogue.xlsx. Also downloads one image per card (large +
small) shared across all variants.

Output workbook sheets:
  Sets  — one row per set (summary), formatted as an Excel Table
  Cards — one row per card, every available field from TCGdex (master
          reference), formatted as an Excel Table

Slice views (e.g. Catalogue for listings, Website export) are built as
Power Query queries reading from the Cards table — no script changes
needed to add or remove columns from a view.

Supports incremental builds — if catalogue.xlsx already exists, new sets
are appended and already-processed sets are skipped automatically.

Usage:
    uv run python scripts/build-catalogue.py              # all in-scope sets
    uv run python scripts/build-catalogue.py --set sv10   # single set
    uv run python scripts/build-catalogue.py --set sv10 --dry-run
"""

import sys
import requests
import argparse
import os
import re
import time

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table

# --- Paths ---
INPUT_PATH  = os.path.join("data", "input", "reference.xlsx")
OUTPUT_PATH = os.path.join("data", "output", "catalogue.xlsx")
IMG_DIR     = os.path.join("data", "output", "images")

# --- APIs ---
TCGDEX_URL      = "https://api.tcgdex.net/v2/en"
FRANKFURTER_URL = "https://api.frankfurter.app/latest?from=USD&to=AUD"

# --- Styling (soft UI — Linear/Stripe aesthetic) ---
ACCENT_COLOR      = "5C5BDB"
HEADER_FILL       = PatternFill("solid", start_color=ACCENT_COLOR)
HEADER_FONT       = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL          = PatternFill("solid", start_color="F4F3FF")
NORMAL_FONT       = Font(name="Calibri", size=10)
HEADER_ROW_HEIGHT = 26
BODY_ROW_HEIGHT   = 16


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fetch(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"    [ERROR] {url}: {e}")
                return None


def slugify(name):
    """Convert a name to a filename-safe slug."""
    name = name.lower()
    name = re.sub(r"['\u2019\u2018`]", "", name)  # strip apostrophes
    name = re.sub(r"[^a-z0-9]+", "-", name)        # non-alphanumeric → hyphen
    return name.strip("-")


def _download_file(url, filepath, retries=3):
    """Download a single file if it doesn't already exist."""
    if os.path.exists(filepath):
        return True
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=20, stream=True)
            r.raise_for_status()
            with open(filepath, "wb") as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            return True
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"    [ERROR] Image download failed: {os.path.basename(filepath)} — {e}")
                return False


def download_images(image_base, series_id, set_id, local_id, card_name):
    """
    Download large (high.jpg) and small (low.jpg) card images.
    Stored at: images/{large|small}/{series_id}/{set_id}/{local_id}_{slug}.jpg
    Returns (large_rel_path, small_rel_path) as forward-slash relative paths.
    """
    if not image_base:
        return "", ""

    fname      = f"{local_id}_{slugify(card_name)}.jpg"
    large_path = os.path.join(IMG_DIR, "large", series_id, set_id, fname)
    small_path = os.path.join(IMG_DIR, "small", series_id, set_id, fname)

    _download_file(f"{image_base}/high.jpg", large_path)
    _download_file(f"{image_base}/low.jpg",  small_path)

    return large_path.replace("\\", "/"), small_path.replace("\\", "/")


def get_usd_to_aud():
    data = fetch(FRANKFURTER_URL)
    if data:
        rate = data.get("rates", {}).get("AUD")
        if rate:
            print(f"  USD->AUD rate: {rate}")
            return rate
    print("  [WARN] Could not fetch exchange rate — price_aud_converted will be empty")
    return None


# ---------------------------------------------------------------------------
# Stringify helpers for complex API fields
# ---------------------------------------------------------------------------

def _str_attacks(attacks):
    if not attacks:
        return ""
    parts = []
    for a in attacks:
        cost   = ",".join(a.get("cost") or [])
        name   = a.get("name", "")
        damage = a.get("damage", "")
        effect = a.get("effect", "")
        part   = name
        if cost:
            part += f" [{cost}]"
        if damage:
            part += f" {damage}"
        if effect:
            part += f" — {effect}"
        parts.append(part)
    return " | ".join(parts)


def _str_abilities(abilities):
    if not abilities:
        return ""
    parts = []
    for a in abilities:
        atype  = a.get("type", "")
        name   = a.get("name", "")
        effect = a.get("effect", "")
        part   = f"{atype}: {name}" if atype else name
        if effect:
            part += f" — {effect}"
        parts.append(part)
    return " | ".join(parts)


def _str_weakness_resistance(items):
    if not items:
        return ""
    return ", ".join(f"{w.get('type', '')} {w.get('value', '')}" for w in items)


def _str_list(lst):
    if not lst:
        return ""
    return ", ".join(str(x) for x in lst)


# ---------------------------------------------------------------------------
# Read input workbook
# ---------------------------------------------------------------------------

def load_sets(wb, target_set_id=None):
    ws = wb["Sets"]
    sets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        set_id, set_name, series_id, series_name, set_type, in_scope, notes = row
        if in_scope != "Yes":
            continue
        if target_set_id and set_id != target_set_id:
            continue
        sets[set_id] = {
            "set_id":      set_id,
            "set_name":    set_name,
            "series_id":   series_id,
            "series_name": series_name,
            "set_type":    set_type,
        }
    return sets


def load_rarities(wb):
    ws = wb["Rarities"]
    rarities = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rarity, base_finish, can_reverse, *_ = row
        rarities[rarity] = {
            "base_finish":      base_finish,
            "can_reverse_holo": (can_reverse == "Yes"),
        }
    return rarities


def load_overrides(wb):
    ws = wb["VariantOverrides"]
    overrides = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        set_id, local_id, card_name, variants_str, reviewed_by, reviewed_date, notes = row
        if not reviewed_by or not reviewed_date:
            continue
        key = (set_id, str(local_id))
        overrides[key] = [v.strip() for v in variants_str.split("|") if v.strip()]
    return overrides


# ---------------------------------------------------------------------------
# Variant logic
# ---------------------------------------------------------------------------

def get_variants_for_card(card, set_info, rarities, overrides):
    set_id   = set_info["set_id"]
    set_type = set_info["set_type"]
    local_id = card.get("localId", "")
    rarity   = card.get("rarity", "") or ""
    tcg_vars = card.get("variants", {}) or {}

    if set_type == "special":
        key = (set_id, local_id)
        if key not in overrides:
            raise ValueError(
                f"Special set {set_id} card {local_id} has no reviewed override entry. "
                f"Run prep-special-set.py and complete the VariantOverrides sheet first."
            )
        return overrides[key]

    rarity_rule = rarities.get(rarity, {"base_finish": "Holo", "can_reverse_holo": False})
    base_finish = rarity_rule["base_finish"]
    can_reverse = rarity_rule["can_reverse_holo"]

    variants = [base_finish]
    if tcg_vars.get("firstEdition"):
        variants.append("First Edition")
    if tcg_vars.get("wPromo"):
        variants.append("Promo")
    if can_reverse and tcg_vars.get("reverse"):
        variants.append("Reverse Holo")

    return variants


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _write_header(ws, headers, col_widths):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def _add_table(ws, table_name):
    """Add a named Excel Table covering just the header row for now.
    Ref is updated to full data range before saving."""
    tab = Table(displayName=table_name, ref=ws.dimensions)
    # No TableStyleInfo — we apply our own cell-level styling
    ws.add_table(tab)


def _update_table_ref(ws, table_name):
    """Extend the named table to cover all current rows."""
    if table_name in ws.tables:
        ws.tables[table_name].ref = ws.dimensions


def _style_row(ws, row_num, values):
    fill = ALT_FILL if row_num % 2 == 0 else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="center", indent=1)
        if fill:
            cell.fill = fill
    ws.row_dimensions[row_num].height = BODY_ROW_HEIGHT


def setup_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # Sets sheet
    ws_sets = wb.create_sheet("Sets")
    _write_header(ws_sets,
                  ["Set ID", "Set Name", "Series ID", "Series Name", "Set Type",
                   "Total Cards", "Release Date"],
                  [14, 28, 12, 20, 12, 14, 14])
    _add_table(ws_sets, "Sets")

    # Cards sheet — master reference, one row per card
    ws_cards = wb.create_sheet("Cards")
    _write_header(ws_cards, [
        "Card ID", "Set ID", "Local ID", "Name", "Category", "Rarity",
        "HP", "Types", "Stage", "Evolves From", "Pokédex ID", "Retreat Cost",
        "Trainer Type", "Abilities", "Attacks", "Weaknesses", "Resistances", "Card Effect",
        "Regulation Mark", "Legal (Standard)", "Legal (Expanded)",
        "Illustrator", "Image (API URL)", "Image (Large)", "Image (Small)",
        "Variants", "Price USD (TCGPlayer)", "Price EUR (Cardmarket)", "Last Updated (API)",
    ], [
        18, 12, 10, 28, 12, 26,
        6, 14, 12, 20, 12, 12,
        14, 40, 60, 20, 20, 40,
        16, 16, 16,
        24, 50, 40, 40,
        30, 20, 20, 22,
    ])
    _add_table(ws_cards, "Cards")

    return wb, ws_sets, ws_cards


def open_or_create_workbook():
    if os.path.exists(OUTPUT_PATH):
        wb = load_workbook(OUTPUT_PATH)
        return wb, wb["Sets"], wb["Cards"], True
    wb, ws_sets, ws_cards = setup_workbook()
    return wb, ws_sets, ws_cards, False


def get_existing_set_ids(ws_sets):
    existing = set()
    for row in ws_sets.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.add(row[0])
    return existing


def write_set_row(ws, row_num, set_info, total_cards, release_date):
    _style_row(ws, row_num, [
        set_info["set_id"], set_info["set_name"],
        set_info["series_id"], set_info["series_name"],
        set_info["set_type"], total_cards, release_date,
    ])


def write_card_row(ws, row_num, card, set_info, variant_labels, img_large, img_small, usd_price, usd_to_aud):
    legal    = card.get("legal") or {}
    pricing  = card.get("pricing") or {}
    cm       = pricing.get("cardmarket") or {}
    eur_price = cm.get("trend") or cm.get("avg")
    aud_price = round(usd_price * usd_to_aud, 2) if usd_price and usd_to_aud else ""

    # Variant summary: e.g. "Normal | Reverse Holo" — useful for quick reference in master
    variants_summary = " | ".join(variant_labels)

    _style_row(ws, row_num, [
        card.get("id", ""),
        set_info["set_id"],
        card.get("localId", ""),
        card.get("name", ""),
        card.get("category", ""),
        card.get("rarity", ""),
        card.get("hp"),
        _str_list(card.get("types")),
        card.get("stage", ""),
        card.get("evolveFrom", ""),
        _str_list(card.get("dexId")),
        card.get("retreat"),
        card.get("trainerType", ""),
        _str_abilities(card.get("abilities")),
        _str_attacks(card.get("attacks")),
        _str_weakness_resistance(card.get("weaknesses")),
        _str_weakness_resistance(card.get("resistances")),
        card.get("effect", ""),
        card.get("regulationMark", ""),
        "Yes" if legal.get("standard") else "No",
        "Yes" if legal.get("expanded") else "No",
        card.get("illustrator", ""),
        card.get("image", ""),
        img_large,
        img_small,
        variants_summary,
        usd_price if usd_price else "",
        eur_price if eur_price else "",
        card.get("updated", ""),
    ])


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_set(set_info, rarities, overrides, ws_sets, ws_cards,
                sets_row, cards_row, usd_to_aud, dry_run):

    set_id    = set_info["set_id"]
    set_name  = set_info["set_name"]
    set_type  = set_info["set_type"]
    series_id = set_info["series_id"]

    print(f"\n{'='*60}")
    print(f"  {set_name} ({set_id}) [{set_type}]")
    print(f"{'='*60}")

    if set_type == "special":
        set_override_keys = [k for k in overrides if k[0] == set_id]
        if not set_override_keys:
            print(f"  [SKIP] Special set {set_id} has no reviewed VariantOverrides entries.")
            print(f"         Run prep-special-set.py and complete the override sheet first.")
            return sets_row, cards_row

    set_data = fetch(f"{TCGDEX_URL}/sets/{set_id}")
    if not set_data:
        print(f"  [ERROR] Could not fetch set data.")
        return sets_row, cards_row

    card_stubs   = set_data.get("cards", [])
    release_date = set_data.get("releaseDate", "")
    total_cards  = len(card_stubs)

    print(f"  {total_cards} cards | Released: {release_date}")

    if not dry_run:
        write_set_row(ws_sets, sets_row, set_info, total_cards, release_date)
    sets_row += 1

    errors    = []
    set_start = time.time()

    for i, stub in enumerate(card_stubs, 1):
        local_id  = stub.get("localId", "")
        card_name = stub.get("name", local_id)

        print(f"  [{i}/{total_cards}] {card_name} ({set_id}-{local_id})", end="")

        card = fetch(f"{TCGDEX_URL}/sets/{set_id}/{local_id}")
        if not card:
            print(f" [ERROR]")
            errors.append(f"{set_id}/{local_id}")
            continue

        try:
            variant_labels = get_variants_for_card(card, set_info, rarities, overrides)
        except ValueError as e:
            print(f" [ERROR] {e}")
            errors.append(f"{set_id}/{local_id}")
            continue

        elapsed   = time.time() - set_start
        rate      = i / elapsed
        remaining = (total_cards - i) / rate if rate > 0 else 0
        eta       = f"{int(remaining // 60)}m{int(remaining % 60):02d}s" if remaining >= 5 else "almost done"
        print(f" -> {', '.join(variant_labels)}  [ETA {eta}]")

        # Download images once per card, shared across all variants
        img_large, img_small = "", ""
        if card.get("image") and not dry_run:
            img_large, img_small = download_images(
                card["image"], series_id, set_id, local_id, card.get("name", local_id)
            )

        # Pricing
        pricing        = (card.get("pricing") or {}).get("tcgplayer") or {}
        normal_pricing = pricing.get("normal") or {}
        usd_price      = normal_pricing.get("marketPrice") or normal_pricing.get("midPrice")

        if not dry_run:
            write_card_row(ws_cards, cards_row, card, set_info,
                           variant_labels, img_large, img_small, usd_price, usd_to_aud)
        cards_row += 1

        time.sleep(0.05)

    set_elapsed  = time.time() - set_start
    rate_per_min = (total_cards / set_elapsed * 60) if set_elapsed > 0 else 0
    print(f"\n  Done: {total_cards} cards, {len(errors)} errors | "
          f"{int(set_elapsed // 60)}m{int(set_elapsed % 60):02d}s | {rate_per_min:.1f} cards/min")
    if errors:
        for e in errors:
            print(f"    - {e}")

    return sets_row, cards_row


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--set", default=None, help="Process a single set ID only")
    parser.add_argument("--dry-run", action="store_true",
                        help="Fetch and print without writing output files")
    args = parser.parse_args()

    if not os.path.exists(INPUT_PATH):
        print(f"Input workbook not found: {INPUT_PATH}")
        print("Run scripts/create-input-workbook.py first.")
        return

    run_start = time.time()
    print("Loading reference workbook...")
    ref_wb    = load_workbook(INPUT_PATH)
    sets      = load_sets(ref_wb, target_set_id=args.set)
    rarities  = load_rarities(ref_wb)
    overrides = load_overrides(ref_wb)

    if not sets:
        msg = f"Set '{args.set}' not found or not in scope." if args.set else "No in-scope sets found."
        print(msg)
        return

    print(f"  {len(sets)} set(s) to process")
    print(f"  {len(rarities)} rarities loaded")
    print(f"  {len(overrides)} override entries loaded")

    print("\nFetching USD->AUD exchange rate...")
    usd_to_aud = get_usd_to_aud()

    dry_run = args.dry_run
    if not dry_run:
        os.makedirs(os.path.join("data", "output"), exist_ok=True)
        wb, ws_sets, ws_cards, is_existing = open_or_create_workbook()
        existing_set_ids = get_existing_set_ids(ws_sets) if is_existing else set()
        sets_row  = ws_sets.max_row + 1
        cards_row = ws_cards.max_row + 1
        if existing_set_ids:
            print(f"\n  Existing catalogue — {len(existing_set_ids)} set(s) already present: "
                  f"{', '.join(sorted(existing_set_ids))}")
    else:
        wb = ws_sets = ws_cards = None
        existing_set_ids = set()
        sets_row = cards_row = 2
        print("  [DRY RUN] No files will be written.")

    start_sets_row  = sets_row
    start_cards_row = cards_row

    for set_info in sets.values():
        if set_info["set_id"] in existing_set_ids:
            print(f"\n  [SKIP] {set_info['set_name']} ({set_info['set_id']}) — already in catalogue")
            continue
        sets_row, cards_row = process_set(
            set_info, rarities, overrides,
            ws_sets, ws_cards,
            sets_row, cards_row,
            usd_to_aud, dry_run,
        )

    run_elapsed = time.time() - run_start
    if not dry_run:
        _update_table_ref(ws_sets,  "Sets")
        _update_table_ref(ws_cards, "Cards")
        wb.save(OUTPUT_PATH)
        print(f"\n{'='*60}")
        print(f"Catalogue saved: {OUTPUT_PATH}")
        print(f"  Sets added:  {sets_row  - start_sets_row}")
        print(f"  Cards added: {cards_row - start_cards_row}")
        print(f"  Images:      {IMG_DIR}/")
    print(f"  Total time: {int(run_elapsed // 60)}m{int(run_elapsed % 60):02d}s")


if __name__ == "__main__":
    main()
